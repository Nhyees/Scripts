import os
import re  # Pour nettoyer les noms de fichiers
from tqdm import tqdm  # Importer tqdm pour la barre de progression
import logging
from openpyxl import load_workbook  # Importer pour lire les fichiers Excel
import yt_dlp  # Importer yt_dlp pour télécharger les vidéos

# Dossier où se trouve le script et les fichiers .xlsx
current_folder = os.getcwd()

print(f"\nCoucou ! Je suis un script conçu pour télécharger les vidéos à partir des liens de ton fichier Excel.")
print(f"\nS'il te plaît, suis attentivement les indications ci-dessous (^-^)/")
print("\n* * *")

# Trouver tous les fichiers .xlsx dans le dossier courant
xlsx_files = [file for file in os.listdir(current_folder) if file.endswith('.xlsx')]

# Vérifier le nbr de fichiers .xlsx
if len(xlsx_files) == 0:
    raise FileNotFoundError("Aucun fichier .xlsx trouvé dans le dossier actuel.")
elif len(xlsx_files) == 1:
    # Un seul fichier trouvé
    xlsx_file = xlsx_files[0]
    print(f"\nFichier Excel (.xlsx) sélectionné : {xlsx_file}")
else:
    # Afficher les fichiers disponibles et demander à l'utilisateur de choisir
    print("\nFichiers Excel (.xlsx) trouvés :")
    for i, file in enumerate(xlsx_files, start=1):
        print(f"{i}: {file}")
    
    # Boucle pour demander un numéro valide
    while True:
        try:
            choice = int(input("\nIndique le numéro du fichier à utiliser : "))
            if 1 <= choice <= len(xlsx_files):
                xlsx_file = xlsx_files[choice - 1]
                break
            else:
                print("S'il te plaît... indique un numéro valide (è_é)")
        except ValueError:
            print("S'il te plaît... indique un numéro valide (è_é)")

    print(f"\nFichier Excel (.xlsx) sélectionné : {xlsx_file}")

# Demander une validation avant de continuer
confirmation = input("\nTape 'OK' pour continuer, ou appuie sur Entrée pour annuler : ").strip()

if confirmation.upper() != 'OK':
    print("Exécution annulée.")
    exit()  # Arrêter l'exécution du script si l'utilisateur n'entre pas 'OK'

# Demander le nom du dossier de sortie
output_folder = input("\nEntre le nom du dossier où les vidéos seront téléchargées (>.<) : ").strip() or "videos"
os.makedirs(output_folder, exist_ok=True)

# Lire les liens et noms à partir du fichier Excel
wb = load_workbook(xlsx_file)
ws = wb.active  # Utiliser la première feuille active

# Boucle pour demander des lettres de colonnes valides
def get_valid_column_letter(prompt):
    while True:
        letter = input(prompt).strip().upper()
        if len(letter) == 1 and letter.isalpha():
            return letter
        else:
            print("S'il te plaît... fais un effort et indique UNE lettre alphabétique (è_é)")

name_column_letter = get_valid_column_letter("\nEntre la lettre de la colonne contenant les titres d'animes (ex: B ou C) : ")
extra_info_column_letter = get_valid_column_letter("Entre la lettre de la colonne contenant les titres de musique (ex: D ou E) : ")
link_column_letter = get_valid_column_letter("Entre la lettre de la colonne contenant les hyperliens à télécharger (ex: D ou E) : ")

# Convertir les lettres en indices
name_column_index = ord(name_column_letter) - ord('A')
extra_info_column_index = ord(extra_info_column_letter) - ord('A')
link_column_index = ord(link_column_letter) - ord('A')

# Fonction pour nettoyer les noms de fichiers
def clean_filename(filename):
    # Remplacer les caractères interdits par un underscore et enlever les sauts de ligne
    filename = re.sub(r'[\/:*?<>|]', '_', filename)  # Remplacer les caractères interdits sauf les guillemets
    filename = filename.replace('"', '')  # Supprimer les guillemets
    filename = filename.replace('\n', ' ').strip()  # Remplacer les sauts de ligne par un espace
    filename = re.sub(r'\s+', ' ', filename)  # Remplacer les espaces multiples par un seul espace
    return filename

# Prétraiter les noms de fichiers
video_names = []
video_links = []

for row in ws.iter_rows(min_row=2):  # Commencer à la ligne 2
    if len(row) > max(name_column_index, extra_info_column_index, link_column_index):  # S'assurer qu'il y ait assez de colonnes
        title = row[name_column_index].value
        extra_info = row[extra_info_column_index].value
        link = row[link_column_index].hyperlink.target if row[link_column_index].hyperlink else None

        if title and link:  # Vérifier que le titre et le lien existent
            video_names.append(clean_filename(f"{title} ({extra_info})"))  # Format: C (D)
            video_links.append(link)

print("\n* * *")
# Compteur pour le nombre total de vidéos
total_videos = len(video_names)
print(f"\nNombre total de vidéos à traiter : {total_videos}")
print("\n* * *")

# Télécharger chaque vidéo à partir des liens et nommer les fichiers selon les noms
for index in range(total_videos):
    video_name = video_names[index]
    link = video_links[index]

    print(f"Traitement de la vidéo {index + 1}/{total_videos}: '{video_name}'")
    
    # Générer le chemin de fichier pour le fichier .mp4
    mp4_file_name = os.path.join(output_folder, f"{video_name}.mp4")

    # Vérifier si le fichier MP4 existe déjà
    if os.path.exists(mp4_file_name):
        print("Le fichier existe déjà, passage à la vidéo suivante.\n")
        print("* * *")
        continue  # Passer à la vidéo suivante

    try:
        # Télécharger la vidéo avec yt-dlp
        ydl_opts = {
            'format': 'bestvideo+bestaudio/best',  # Meilleure qualité vidéo et audio
            'outtmpl': mp4_file_name  # Spécifier le chemin de sortie
        }
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            ydl.download([link])  # Télécharger la vidéo

        print(f"Vidéo sauvegardée dans {output_folder} en mp4.\n")

    except Exception as e:
        print(f"Erreur lors du téléchargement pour {link}: {e}\n")

    print("* * *")  # Étoile après chaque vidéo traitée >.<

print("\nTéléchargement terminé ! Courage pour la suite (>.<)")
